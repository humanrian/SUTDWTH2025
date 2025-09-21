import os
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
import openpyxl
import serial
import time
from twilio.rest import Client
from apscheduler.schedulers.background import BackgroundScheduler
import atexit
from datetime import datetime, timedelta
import asyncio
from notificationapi_python_server_sdk import notificationapi

from credentials import (
    LOGIN_USERNAME,
    LOGIN_PASSWORD,
    SECRET_KEY,
    SECRET_KEY_FILE,
    TWILIO_ACCOUNT_SID,
    TWILIO_AUTH_TOKEN,
    TWILIO_NUMBER,
    RECIPIENT_NUMBER,
    ARDUINO_PORT,
    CALL_LANGUAGE,
    PATIENT_USERNAME,
    PATIENT_PASSWORD,
    NOTIFICATIONAPI1D,
    NOTIFICATIONAPI2D,
    NOTIFICATIONAPIID,
    CARE_NUMBER
)

notificationapi.init(NOTIFICATIONAPI1D, NOTIFICATIONAPI2D)

def get_secret_key():
    if os.path.exists(SECRET_KEY_FILE):
        with open(SECRET_KEY_FILE, 'r') as f:
            return f.read().strip()
    else:
        new_key = os.urandom(64).hex()
        with open(SECRET_KEY_FILE, 'w') as f:
            f.write(new_key)
        return new_key

app = Flask(__name__)

app.secret_key = get_secret_key()

client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
twilio_number = TWILIO_NUMBER
recipient_number = RECIPIENT_NUMBER
call_language = CALL_LANGUAGE

ARDUINO_PORT = ARDUINO_PORT
BAUD_RATE = 115200

def get_sheet():
    try:
        wb = openpyxl.load_workbook("data.xlsx")
        sheet = wb.active
        
        headers = [cell.value for cell in sheet[1]]
        if "Container" not in headers:
            sheet.cell(row=1, column=5, value="Container")
            wb.save("data.xlsx")
            
        return wb, sheet
    except FileNotFoundError:
        return None, None

def update_credentials_file(data):
    """Writes the updated credentials to the credentials.py file."""
    with open('credentials.py', 'w') as f:
        f.write(f"LOGIN_USERNAME = \"{data['username']}\"\n")
        f.write(f"LOGIN_PASSWORD = \"{data['password']}\"\n")
        f.write(f"PATIENT_USERNAME = \"{data['patient_username']}\"\n")
        f.write(f"PATIENT_PASSWORD = \"{data['patient_password']}\"\n")
        f.write(f"SECRET_KEY = \"{get_secret_key()}\"\n")
        f.write("SECRET_KEY_FILE = 'secret_key.txt'\n")
        f.write(f"TWILIO_ACCOUNT_SID = \"{data['twilio_sid']}\"\n")
        f.write(f"TWILIO_AUTH_TOKEN = \"{data['twilio_token']}\"\n")
        f.write(f"TWILIO_NUMBER = \"{data['twilio_number']}\"\n")
        f.write(f"RECIPIENT_NUMBER = \"{data['recipient_number']}\"\n")
        f.write(f"CARE_NUMBER = \"{data['care_number']}\"\n")
        f.write(f"ARDUINO_PORT = \"{data['arduino_port']}\"\n")
        f.write(f"CALL_LANGUAGE = \"{data['call_language']}\"\n")
        f.write(f"NOTIFICATIONAPI1D = \"{NOTIFICATIONAPI1D}\"\n")
        f.write(f"NOTIFICATIONAPI2D = \"{NOTIFICATIONAPI2D}\"\n")
        f.write(f"NOTIFICATIONAPIID = \"{NOTIFICATIONAPIID}\"\n")

def dispense_medication_job():
    print("Checking for scheduled medication at:", datetime.now().strftime("%H:%M"))
    wb, sheet = get_sheet()
    if not sheet:
        print("Error: data.xlsx not found.")
        return
    
    current_time_str = datetime.now().strftime("%H:%M")
    
    meds_to_dispense = []
    for row_num in range(2, sheet.max_row + 1):
        med_time = str(sheet.cell(row=row_num, column=1).value)
        if med_time == current_time_str:
            meds_to_dispense.append(sheet.cell(row=row_num, column=2).value)

    if meds_to_dispense:
        print("Scheduled dispensing triggered for:", current_time_str)
        
        if CALL_LANGUAGE == "English":
            twiml_message = '<Response><Say>Medication time, please take your meds. Medication time, please take your meds.Medication time, please take your meds. Medication time, please take your meds.</Say></Response>'
        elif CALL_LANGUAGE == "Chinese":
            twiml_message = '<Response><Say language="zh-CN">服药时间到了，请吃药.服药时间到了，请吃药.服药时间到了，请吃药.</Say></Response>'
        else:
            twiml_message = '<Response><Say>Medication time, please take your meds.</Say></Response>'

        try:
            call = client.calls.create(
                twiml=twiml_message,
                to=recipient_number,
                from_=twilio_number
            )
            print("Call ID:", call.sid)

            async def send_notification():
                response = await notificationapi.send({
                "type": "medication",
                "to": {
                    "id": NOTIFICATIONAPIID,
                    "number": CARE_NUMBER
                },
                    "sms": {
                        "message": "Medication notification sent"
                    }
                })

            asyncio.run(send_notification())

            time.sleep(2)

        except Exception as e:
            print("Error making the call:", str(e))

@app.route('/')
def home():
    if not session.get('logged_in'):
        return redirect(url_for('landing_page'))
    if session.get('user_role') == 'patient':
        return redirect(url_for('patient_dashboard'))
    return redirect(url_for('index'))

@app.route('/index')
def index():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    if session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('patient_dashboard'))
    return render_template('index.html')

@app.route('/landing_page')
def landing_page():
    return render_template('landing.html')

@app.route('/patient_dashboard')
def patient_dashboard():
    if not session.get('logged_in'):
        flash('Please log in to access this page.')
        return redirect(url_for('login'))
    
    if session.get('user_role') != 'patient':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))

    wb, sheet = get_sheet()
    if not sheet:
        flash("Error: data.xlsx not found.")
        return redirect(url_for('index'))

    all_data = []
    meds_to_take_now = []
    
    current_time = datetime.now()
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1]:
            all_data.append(row[:4])
            
            try:
                med_time_str = str(row[0])
                med_hour = int(med_time_str.split(':')[0])
                med_minute = int(med_time_str.split(':')[1])
                med_datetime = current_time.replace(hour=med_hour, minute=med_minute, second=0, microsecond=0)
                
                time_difference = abs(current_time - med_datetime)
                if time_difference <= timedelta(minutes=5):
                    meds_to_take_now.append(row[1])
            except (ValueError, IndexError):
                continue
                
    return render_template('patient_dashboard.html', data=all_data, meds_to_take_now=meds_to_take_now)

@app.route('/taken_medication', methods=['POST'])
def taken_medication():
    print(f"Medication has been marked as taken.")
    
    async def send_notification():
        response = await notificationapi.send({
        "type": "medication",
        "to": {
            "id": NOTIFICATIONAPIID,
            "number": CARE_NUMBER
        },
            "sms": {
                "message": "Medication has been taken"
            }
        })

    asyncio.run(send_notification())
    
    return jsonify(success=True)

@app.route('/show_all')
def show_all():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))

    wb, sheet = get_sheet()
    if not sheet:
        flash("Error: data.xlsx not found.")
        return redirect(url_for('index'))
    
    all_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1]:
            all_data.append(row[:4])
            
    return render_template('show_all.html', data=all_data)

@app.route('/login')
def login():
    if session.get('logged_in'):
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/login_attempt', methods=['POST'])
def login_attempt():
    username = request.form['username']
    password = request.form['password']

    if username == LOGIN_USERNAME and password == LOGIN_PASSWORD:
        session['logged_in'] = True
        session['user_role'] = 'caregiver'
        flash('Login successful!')
        return redirect(url_for('index'))
    elif username == PATIENT_USERNAME and password == PATIENT_PASSWORD:
        session['logged_in'] = True
        session['user_role'] = 'patient'
        flash('Login successful!')
        return redirect(url_for('patient_dashboard'))
    else:
        flash('Invalid username or password.')
        return redirect(url_for('login'))

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    session.pop('user_role', None)
    flash('You have been logged out.')
    return redirect(url_for('login'))

@app.route('/edit')
def edit_data():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))

    wb, sheet = get_sheet()
    if not sheet:
        flash("Error: data.xlsx not found.")
        return redirect(url_for('edit_data'))
    
    meds_existing = [sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=2).value]
    
    used_containers = [sheet.cell(row=i, column=4).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=4).value]
    
    all_containers = set(range(1, 11))
    used_containers_set = {int(c) for c in used_containers if c and str(c).isdigit()}
    available_containers = sorted(list(all_containers - used_containers_set))
    
    return render_template('edit_data.html', meds_existing=meds_existing, available_containers=available_containers)

@app.route('/config')
def config():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))
        
    return render_template('config.html', 
        LOGIN_USERNAME=LOGIN_USERNAME,
        LOGIN_PASSWORD=LOGIN_PASSWORD,
        PATIENT_USERNAME=PATIENT_USERNAME,
        PATIENT_PASSWORD=PATIENT_PASSWORD,
        TWILIO_ACCOUNT_SID=TWILIO_ACCOUNT_SID,
        TWILIO_AUTH_TOKEN=TWILIO_AUTH_TOKEN,
        TWILIO_NUMBER=TWILIO_NUMBER,
        CARE_NUMBER=CARE_NUMBER,
        RECIPIENT_NUMBER=RECIPIENT_NUMBER,
        ARDUINO_PORT=ARDUINO_PORT,
        CALL_LANGUAGE=CALL_LANGUAGE,
        NOTIFICATIONAPI1D=NOTIFICATIONAPI1D,
        NOTIFICATIONAPI2D=NOTIFICATIONAPI2D,
        NOTIFICATIONAPIID=NOTIFICATIONAPIID
    )

@app.route('/save_config', methods=['POST'])
def save_config():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))
        
    new_credentials = {
        'username': request.form['username'],
        'password': request.form['password'],
        'patient_username': request.form['patient_username'],
        'patient_password': request.form['patient_password'],
        'twilio_sid': request.form['twilio_sid'],
        'twilio_token': request.form['twilio_token'],
        'twilio_number': request.form['twilio_number'],
        'recipient_number': request.form['recipient_number'],
        'care_number': request.form['care_number'],
        'arduino_port': request.form['arduino_port'],
        'call_language': request.form['call_language'],
        'notificationapi1d': request.form['notificationapi1d'],
        'notificationapi2d': request.form['notificationapi2d'],
        'notificationapiid': request.form['notificationapiid']
    }
    
    update_credentials_file(new_credentials)

    session.pop('logged_in', None)
    flash("Configuration saved successfully! Please restart the Flask server to apply changes.")
    
    return redirect(url_for('login'))

@app.route('/confirm_add', methods=['POST'])
def confirm_add():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))
    
    time = request.form['med_time']
    name = request.form['med_name']
    amount = request.form['med_amount']
    container = request.form['container']

    return render_template('confirm_add.html', time=time, name=name, amount=amount, container=container)

@app.route('/do_add_med', methods=['POST'])
def do_add_med():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))

    wb, sheet = get_sheet()
    if not sheet:
        flash("Error: data.xlsx not found.")
        return redirect(url_for('edit_data'))

    med_details = [
        request.form['med_time'],
        request.form['med_name'],
        request.form['med_amount'],
        request.form['container']
    ]
    
    sheet.append(med_details)
    wb.save("data.xlsx")
    flash(f"Medicine '{med_details[1]}' was added successfully!")
    return redirect(url_for('edit_data'))

@app.route('/confirm_edit', methods=['POST'])
def confirm_edit():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))
        
    wb, sheet = get_sheet()
    if not sheet:
        flash("Error: data.xlsx not found.")
        return redirect(url_for('edit_data'))

    med_name_edit = request.form['med_to_edit']
    new_name = request.form['new_name']
    new_time = request.form['new_time']
    new_amount = request.form['new_amount']
    new_container = request.form['new_container']
    
    old_data = []
    for row_num in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_num, column=2).value == med_name_edit:
            old_data.append(sheet.cell(row=row_num, column=1).value)
            old_data.append(sheet.cell(row=row_num, column=2).value)
            old_data.append(sheet.cell(row=row_num, column=3).value)
            old_data.append(sheet.cell(row=row_num, column=4).value)
            break
    
    if old_data:
        return render_template('confirm_edit.html', old_data=old_data, new_name=new_name, new_time=new_time, new_amount=new_amount, new_container=new_container)
    else:
        flash("Error: Medicine not found for editing.")
        return redirect(url_for('edit_data'))

@app.route('/do_edit_med', methods=['POST'])
def do_edit_med():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))
    
    wb, sheet = get_sheet()
    if not sheet:
        flash("Error: data.xlsx not found.")
        return redirect(url_for('edit_data'))

    med_name_edit = request.form['med_to_edit']
    new_name = request.form['new_name']
    new_time = request.form['new_time']
    new_amount = request.form['new_amount']
    new_container = request.form['new_container']

    for row_num in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_num, column=2).value == med_name_edit:
            if new_time:
                sheet.cell(row=row_num, column=1).value = new_time
            if new_name:
                sheet.cell(row=row_num, column=2).value = new_name
            if new_amount:
                sheet.cell(row=row_num, column=3).value = new_amount
            if new_container:
                sheet.cell(row=row_num, column=4).value = new_container
            break
    
    wb.save("data.xlsx")
    flash(f"Medicine '{med_name_edit}' was updated successfully!")
    return redirect(url_for('edit_data'))

@app.route('/confirm_delete', methods=['POST'])
def confirm_delete():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))
        
    wb, sheet = get_sheet()
    if not sheet:
        flash("Error: data.xlsx not found.")
        return redirect(url_for('edit_data'))
    
    med_name_delete = request.form['med_to_delete']
    med_data = []

    for row_num in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_num, column=2).value == med_name_delete:
            med_data.append(sheet.cell(row=row_num, column=1).value)
            med_data.append(sheet.cell(row=row_num, column=2).value)
            med_data.append(sheet.cell(row=row_num, column=3).value)
            med_data.append(sheet.cell(row=row_num, column=4).value)
            break
    
    if med_data:
        return render_template('confirm_delete.html', med_data=med_data)
    else:
        flash("Error: Medicine not found.")
        return redirect(url_for('edit_data'))

@app.route('/delete_med', methods=['POST'])
def delete_med():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))
        
    wb, sheet = get_sheet()
    if not sheet:
        flash("Error: data.xlsx not found.")
        return redirect(url_for('edit_data'))

    med_name_delete = request.form['med_name_delete']
    for row_num in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_num, column=2).value == med_name_delete:
            sheet.delete_rows(row_num, 1)
            break
            
    wb.save("data.xlsx")
    flash(f"Medicine '{med_name_delete}' was deleted successfully!")
    return redirect(url_for('edit_data'))

@app.route('/run')
def run_simulation():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))
        
    wb, sheet = get_sheet()
    if not sheet:
        flash("Error: data.xlsx not found.")
        return redirect(url_for('index'))

    unique_times = sorted(list(set(str(row[0]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[0])))
    
    return render_template('run_simulation.html', unique_times=unique_times)

@app.route('/dispense', methods=['POST'])
def dispense():
    if not session.get('logged_in') or session.get('user_role') != 'caregiver':
        flash('You do not have permission to access this page.')
        return redirect(url_for('index'))
        
    timing = request.form['timing']
    wb, sheet = get_sheet()
    if not sheet:
        flash("Error: data.xlsx not found.")
        return redirect(url_for('run_simulation'))
    
    meds_dispense = {}
    for row_num in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row_num, column=1).value) == timing:
            meds_dispense[sheet.cell(row=row_num, column=2).value] = {
                "quantity": sheet.cell(row=row_num, column=3).value,
                "container": sheet.cell(row=row_num, column=4).value
            }

    print("Meds to dispense:")
    for name, details in meds_dispense.items():
        print(f"Name: {name}, Quantity: {details['quantity']}, Container: {details['container']}")

    if CALL_LANGUAGE == "English":
        twiml_message = '<Response><Say>Medication time, please take your meds. Medication time, please take your meds.Medication time, please take your meds. Medication time, please take your meds.</Say></Response>'
    elif CALL_LANGUAGE == "Chinese":
        twiml_message = '<Response><Say language="zh-CN">服药时间到了，请吃药.服药时间到了，请吃药.服药时间到了，请吃药.</Say></Response>'
    else:
        twiml_message = '<Response><Say>Medication time, please take your meds.</Say></Response>'

    try:
        call = client.calls.create(
            twiml=twiml_message,
            to=recipient_number,
            from_=twilio_number
        )
        print("Call ID:", call.sid)

        async def send_notification():
            response = await notificationapi.send({
            "type": "medication",
            "to": {
                "id": NOTIFICATIONAPIID,
                "number": CARE_NUMBER
            },
                "sms": {
                    "message": "Medication notification sent"
                }
            })

        asyncio.run(send_notification())

        time.sleep(2)

    except Exception as e:
        print("Error making the call:", str(e))
        flash(f"Error making the call: {str(e)}")

    flash(f"Dispensing for {timing} confirmed successfully! Check the console for details.")
    return redirect(url_for('run_simulation'))

if __name__ == '__main__':
    scheduler = BackgroundScheduler()
    scheduler.add_job(func=dispense_medication_job, trigger="interval", minutes=1)
    scheduler.start()

    app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False)

    atexit.register(lambda: scheduler.shutdown())